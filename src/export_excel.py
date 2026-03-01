"""Export all pipeline artifacts into a single multi-sheet Excel workbook.

Each sheet gets auto-fitted column widths, a frozen header row, and styled
headers so the file is immediately browsable in Excel / Google Sheets /
LibreOffice without manual tweaking.
"""

from __future__ import annotations

import json
import logging
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

_SHEETS: list[tuple[str, str, str]] = [
    ("resumes_clean", "parquet", "data.clean_parquet"),
    ("variants", "parquet", "data.variants_parquet"),
    ("variants_scored", "csv", "data.variants_scored"),
    ("eval_detail", "csv", "data.eval_summary_dir/eval_detail.csv"),
    ("eval_summary", "csv", "data.eval_summary_dir/eval_summary.csv"),
    ("predictions", "csv", "logistic_regression.reports_dir/logistic_regression_predictions.csv"),
    ("coefficients", "csv", "logistic_regression.reports_dir/logistic_regression_coefficients.csv"),
    ("lr_metrics", "json", "logistic_regression.reports_dir/logistic_regression_metrics.json"),
]


def _resolve_path(config: dict, spec: str) -> Path:
    """Resolve a dotted config key, optionally appended with a literal filename."""
    parts = spec.split("/", maxsplit=1)
    key_chain = parts[0].split(".")
    val = config
    for k in key_chain:
        val = val[k]
    p = Path(val)
    if len(parts) == 2:
        p = p / parts[1]
    return p


def _read_artifact(path: Path, fmt: str) -> pd.DataFrame | None:
    if not path.exists():
        logger.warning("Artifact not found, skipping: %s", path)
        return None
    if fmt == "parquet":
        return pd.read_parquet(path)
    if fmt == "csv":
        return pd.read_csv(path)
    if fmt == "json":
        with open(path) as f:
            data = json.load(f)
        flat: dict[str, object] = {}
        for k, v in data.items():
            if isinstance(v, dict):
                for sub_k, sub_v in v.items():
                    flat[f"{k}.{sub_k}"] = sub_v
            else:
                flat[k] = v
        return pd.DataFrame([flat])
    return None


def _style_sheet(ws: "openpyxl.worksheet.worksheet.Worksheet") -> None:
    """Freeze top row, style headers, and auto-fit column widths."""
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1), start=1):
        max_len = 0
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except TypeError:
                pass
        adjusted = min(max_len + 3, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def export_excel(config: dict) -> Path:
    """Build the multi-sheet workbook and return its path."""
    out_path = Path(config.get("export", {}).get(
        "excel_path", "outputs/all_artifacts.xlsx",
    ))
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, fmt, spec in _SHEETS:
            path = _resolve_path(config, spec)
            df = _read_artifact(path, fmt)
            if df is None:
                continue
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            logger.info(
                "Sheet '%s' ← %s  (%d × %d)",
                sheet_name, path, len(df), len(df.columns),
            )

        for sheet_name in writer.book.sheetnames:
            _style_sheet(writer.book[sheet_name])

    logger.info("Excel workbook written → %s", out_path)
    return out_path
